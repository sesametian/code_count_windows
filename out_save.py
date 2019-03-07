#encoding=utf-8
import os,time
from openpyxl import Workbook
from openpyxl import load_workbook

def out_to_xls(file_dict):
    nowtime=time.strftime("%Y-%m-%d %H-%M-%S", time.localtime()) #获取当前时间并格式化输出
    file_path=os.path.dirname(__file__)
    if os.path.exists(file_path+'\out_save.xlsx'):
        #print("y")
        wb=load_workbook(file_path+'\out_save.xlsx')
        ws=wb.create_sheet(nowtime)
        ws['A1']='文件名' #增加表头
        ws['B1']='有效代码'
        ws['C1']='空白行数'
        ws['D1']='注释行数'
    
        for file_name,file_data in file_dict.items(): #循环得到的文件字典
            ws.append([file_name]+list(file_data)) #因为工作表的append方法只能添加一个list，所以把文件名和文件统计数据放在一个list里

        wb.save(file_path+'\out_save.xlsx') #把导出的内容保存到文件目录下
    else:
        #print("no")
        wb=Workbook() #创建一个工作簿
        ws=wb.create_sheet(nowtime,index=0) #新建名称为当前时间的sheet插在开头，方便做统计及记录
        ws['A1']='文件名' #增加表头
        ws['B1']='有效代码'
        ws['C1']='空白行数'
        ws['D1']='注释行数'
    
        for file_name,file_data in file_dict.items(): #循环得到的文件字典
            ws.append([file_name]+list(file_data)) #因为工作表的append方法只能添加一个list，所以把文件名和文件统计数据放在一个list里
        
        wb.remove(wb.get_sheet_by_name("Sheet"))
        wb.save(file_path+'\out_save.xlsx') #把导出的内容保存到文件目录下
